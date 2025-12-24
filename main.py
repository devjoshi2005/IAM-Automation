from pydantic import ConfigDict
import os, json, uuid, csv, io
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse
from celery import Celery, Task
from msal import ConfidentialClientApplication
import requests, hvac
from sqlalchemy import create_engine
from sqlmodel import Field, SQLModel, Session, create_engine as create_sqlmodel_engine
from okta.client import Client as OktaClient

# Vault config - deferred loading
vault = None
okta_sec = None
azure_sec = None

def get_vault_secrets():
    """Load secrets from Vault with fallback to environment variables"""
    global vault, okta_sec, azure_sec
    
    if okta_sec and azure_sec:
        return okta_sec, azure_sec
    
    try:
        vault = hvac.Client(url=os.getenv("VAULT_ADDR", "http://localhost:8200"), token=os.getenv("VAULT_TOKEN", "root"))
        okta_sec = vault.secrets.kv.v2.read_secret_version(path="iam/okta")["data"]["data"]
        azure_sec = vault.secrets.kv.v2.read_secret_version(path="iam/azure")["data"]["data"]
        print("✓ Secrets loaded from Vault")
    except Exception as e:
        print(f"⚠ Vault connection failed: {e}. Using environment variables.")
        okta_sec = {
            "org url": os.getenv("OKTA_ORG_URL"),
            "token": os.getenv("OKTA_TOKEN")
        }
        azure_sec = {
            "tenant ID": os.getenv("AZ_TENANT_ID"),
            "client id": os.getenv("AZ_CLIENT_ID"),
            "secret": os.getenv("AZ_SECRET")
        }
        if not all([okta_sec.get("org url"), okta_sec.get("token"), azure_sec.get("tenant ID"), azure_sec.get("client id"), azure_sec.get("secret")]):
            raise Exception("Missing required environment variables or Vault secrets")
    
    return okta_sec, azure_sec

ROLE_OKTA_APPS = {
    "Personal-Assistant": ["google-workspace-calendar", "outlook"],
    "Junior-Developer": ["slack", "teams"],
    "Sales-Rep": ["slack", "teams"]
}
ROLE_AZ_GROUP = {
    "Personal-Assistant": "PA-GROUP",
    "Junior-Developer": "Dev-GROUP",
    "Sales-Rep": "Sales-GROUP"
}

class ProvisionLog(SQLModel, table=True):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    email: str = Field(primary_key=True)
    role: str
    status: str
    timestamp: datetime = Field(default_factory=datetime.now)

engine = create_sqlmodel_engine(os.getenv("DB_URL", "postgresql+psycopg2://postgres:postgres@localhost:5432/iam"), echo=False)
SQLModel.metadata.create_all(engine)

def make_celery(app_name):
    return Celery(app_name, broker="redis://localhost:6379/0", backend="redis://localhost:6379/0")

celery_app = make_celery("tasks")

class MyTask(Task):
    def on_failure(self, exc, task_id, args, kwargs, einfo):
        print(f"Task {task_id} failed: {exc}")

celery_app.Task = MyTask

@celery_app.task(bind=True)
def provision_users(self, records):
    try:
        okta_sec, azure_sec = get_vault_secrets()
        OKTA_DOMAIN = okta_sec["org url"]
        OKTA_TOKEN = okta_sec["token"]
        AZ_TENANT = azure_sec["tenant ID"]
        AZ_CLIENT = azure_sec["client id"]
        AZ_SECRET = azure_sec["secret"]
        
        okta_client = OktaClient({"orgUrl": OKTA_DOMAIN, "token": OKTA_TOKEN})
        msal_app = ConfidentialClientApplication(AZ_CLIENT, authority=f"https://login.microsoftonline.com/{AZ_TENANT}", client_credential=AZ_SECRET)
        token = msal_app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        if "access_token" not in token: 
            raise Exception("Azure token acquisition failed")
        headers = {"Authorization": f"Bearer {token['access_token']}", "Content-Type": "application/json"}

        for rec in records:
            with Session(engine) as session:
                try:
                    # Create Okta user
                    user_payload = {
                        "profile": {
                            "firstName": rec["firstname"],
                            "lastName": rec["lastname"],
                            "email": rec["email"],
                            "login": rec["email"]
                        },
                        "credentials": {
                            "password": {
                                "value": str(uuid.uuid4())
                            }
                        }
                    }
                    okta_user = okta_client.create_user(user_payload, activate=True)
                    
                    # Assign Okta apps
                    for app_name in ROLE_OKTA_APPS.get(rec["role"], []):
                        try:
                            okta_client.assign_user_to_application(okta_user["id"], app_name)
                        except Exception as app_err:
                            print(f"Warning: Could not assign app {app_name}: {app_err}")
                    
                    # Create Azure user
                    user_pay = {
                        "accountEnabled": True,
                        "displayName": f"{rec['firstname']} {rec['lastname']}",
                        "mailNickname": rec["email"].split("@")[0],
                        "userPrincipalName": rec["email"],
                        "passwordProfile": {
                            "password": str(uuid.uuid4()),
                            "forceChangePasswordNextSignIn": True
                        }
                    }
                    az_r = requests.post("https://graph.microsoft.com/v1.0/users", headers=headers, json=user_pay)
                    if az_r.status_code in (200, 201):
                        az_id = az_r.json()["id"]
                        
                        # Add user to Azure group
                        gid_resp = requests.get(
                            f"https://graph.microsoft.com/v1.0/groups?$filter=displayName eq '{ROLE_AZ_GROUP[rec['role']]}'",
                            headers=headers
                        )
                        if gid_resp.status_code == 200 and gid_resp.json().get("value"):
                            gid = gid_resp.json()["value"][0]["id"]
                            requests.post(
                                f"https://graph.microsoft.com/v1.0/groups/{gid}/members/$ref",
                                headers=headers,
                                json={"@odata.id": f"https://graph.microsoft.com/v1.0/directoryObjects/{az_id}"}
                            )
                        else:
                            print(f"Warning: Group not found for role {rec['role']}")
                    else:
                        raise Exception(f"Azure user creation failed: {az_r.text}")
                    
                    # Log success
                    log_entry = ProvisionLog(email=rec["email"], role=rec["role"], status="done")
                    session.add(log_entry)
                    session.commit()
                except Exception as e:
                    log_entry = ProvisionLog(email=rec["email"], role=rec["role"], status=f"error: {str(e)}")
                    session.add(log_entry)
                    session.commit()
        
        return {"total": len(records), "status": "finished"}
    except Exception as e:
        print(f"Task provisioning error: {str(e)}")
        raise

app = FastAPI(title="IAM-PAM Bulk Provision")

@app.get("/", response_class=HTMLResponse)
def index():
    return open("index.html").read()

@app.post("/upload")
def upload(file: UploadFile = File(...)):
    if file.content_type not in {"text/csv", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
        raise HTTPException(400, "CSV/XLSX only")
    if file.filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(file.file.read().decode("utf-8")))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
        headers = [c.value.lower() for c in ws[1]]
        reader = [dict(zip(headers, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]
    required = ["email", "firstname", "lastname", "role", "department", "manageremail", "startdate", "location", "workertype"]
    rows = list(reader)
    if not rows: 
        raise HTTPException(400, "Empty file")
    if list(rows[0].keys()) != required:
        raise HTTPException(400, f"Column mismatch. Expected: {required}")
    bad_roles = set(r["role"] for r in rows) - set(ROLE_OKTA_APPS.keys())
    if bad_roles: 
        raise HTTPException(400, f"Roles not allowed: {bad_roles}")
    task = provision_users.delay(rows)
    return {"task_id": task.id, "message": "File accepted, provisioning started"}

@app.get("/result/{task_id}")
def result(task_id: str):
    res = celery_app.AsyncResult(task_id)
    if res.state == "PENDING": 
        return {"state": "PENDING"}
    if res.state == "FAILURE": 
        return {"state": "FAILURE", "detail": str(res.info)}
    return {"state": res.state, "info": res.result}